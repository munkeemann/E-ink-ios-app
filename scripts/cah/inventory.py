#!/usr/bin/env python3
"""
SAM1-60: inventory helper for the Reddit CAH master spreadsheet.

Two modes:
  - sheets (default): list every sheet with row/col counts + first four rows.
  - sets  <sheet>:    list every unique Set value in the given sheet's
                      left-half (col C) and right-half (col H) with counts
                      and sample card texts.

Expected xlsx path: ~/cah-master.xlsx (Google Sheets export of the Reddit
master CAH spreadsheet, see SAM1-60 ticket for source URL).

Usage:
  python3 scripts/cah/inventory.py sheets
  python3 scripts/cah/inventory.py sets "Master Cards List"
"""

import os
import sys
from collections import defaultdict
from openpyxl import load_workbook

XLSX_PATH = os.path.expanduser('~/cah-master.xlsx')


def truncate(v, n=80):
    s = '' if v is None else str(v)
    s = s.replace('\n', '\\n').replace('\r', '\\r')
    return s if len(s) <= n else s[:n - 1] + '…'


def sheets_report():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f'Workbook: {XLSX_PATH}')
    print(f'Sheet count: {len(wb.sheetnames)}\n')
    for name in wb.sheetnames:
        ws = wb[name]
        total = 0
        non_empty = 0
        max_col = 0
        first_four = []
        for row in ws.iter_rows(values_only=True):
            total += 1
            if any(v is not None and str(v).strip() for v in row):
                non_empty += 1
            if row and len(row) > max_col:
                max_col = len(row)
            if len(first_four) < 4:
                first_four.append(row)
        print(f'\n▸ {name!r}  — total={total}  non_empty={non_empty}  cols={max_col}')
        if first_four:
            print('  r1 header:')
            for ci, v in enumerate(first_four[0] or [], 1):
                if v is not None and str(v).strip():
                    print(f'    col{ci:>2}: {truncate(v, 100)!r}')
            print('  r2..r4 data:')
            for ri, row in enumerate(first_four[1:], 2):
                non = [(ci, v) for ci, v in enumerate(row or [], 1)
                       if v is not None and str(v).strip()]
                if not non:
                    print(f'    r{ri}: <empty>')
                    continue
                cells = ', '.join(f'c{ci}={truncate(v, 60)!r}' for ci, v in non[:4])
                more = '' if len(non) <= 4 else f' (+{len(non) - 4} more cols)'
                print(f'    r{ri}: {cells}{more}')
    wb.close()


def sets_report(sheet_name: str):
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f'ERROR: sheet {sheet_name!r} not found. Available:', file=sys.stderr)
        for n in wb.sheetnames:
            print(f'  {n}', file=sys.stderr)
        wb.close()
        sys.exit(2)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    print(f'Sheet: {sheet_name!r}')
    print(f'r1 header: {rows[0]!r}')

    left = defaultdict(list)
    right = defaultdict(list)
    # Assumed layout (matches Master Cards List):
    #   col A=0 text, col C=2 Set (left);  col G=6 text, col H=7 Set (right)
    for row in rows[1:]:
        if len(row) >= 3:
            t, s = row[0], row[2]
            if t and str(t).strip() and s and str(s).strip():
                left[str(s).strip()].append(str(t).strip())
        if len(row) >= 8:
            t, s = row[6], row[7]
            if t and str(t).strip() and s and str(s).strip():
                right[str(s).strip()].append(str(t).strip())

    def dump(label, table):
        sorted_ = sorted(table.items(), key=lambda x: (-len(x[1]), x[0]))
        total = sum(len(v) for v in table.values())
        print(f'\n── {label}  ({len(sorted_)} unique, {total} rows) ──')
        for i, (name, texts) in enumerate(sorted_, 1):
            samples = ' | '.join(truncate(t, 50) for t in texts[:3])
            print(f'{i:>4}  {len(texts):>6}  {name}')
            if samples:
                print(f'         {samples}')

    dump('LEFT-half Set column (prompts side, col C)', left)
    dump('RIGHT-half Set column (responses side, col H)', right)


def main():
    args = sys.argv[1:]
    if not args or args[0] == 'sheets':
        sheets_report()
    elif args[0] == 'sets' and len(args) >= 2:
        sets_report(' '.join(args[1:]))
    else:
        print(__doc__, file=sys.stderr)
        sys.exit(2)


if __name__ == '__main__':
    main()
