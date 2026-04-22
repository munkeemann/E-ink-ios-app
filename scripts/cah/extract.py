#!/usr/bin/env python3
"""
SAM1-60: extract CAH-prefix sets from the Reddit master spreadsheet to
src/assets/data/cah_content.json.

Source: Master Cards List sheet of ~/cah-master.xlsx. Two-half row layout:
  cols A-E: Prompt Cards | Special | Set | Sheet | Version Comments
  col  F  : blank separator
  cols G-J: Response Cards | Set | Sheet | Version Comments

Decision rule: include any Set whose normalized name starts with 'CAH '.
Normalization applies a hand-curated rename map (seven known left/right
spelling mismatches) and a defensive whitespace strip before filtering.

Pick-count derivation:
  - If Special col contains 'PICK N' (case-insensitive), use N.
  - Otherwise count runs of 2+ underscores in the prompt text:
    3+ blanks → pick 3; 2 blanks → pick 2; else pick 1.

Per-pack exact-text dedup on both sides (same card listed twice in the
merged list → kept once).

Usage: python3 scripts/cah/extract.py
"""

import hashlib
import json
import os
import re
from collections import defaultdict
from openpyxl import load_workbook

XLSX_PATH = os.path.expanduser('~/cah-master.xlsx')
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
OUT_PATH = os.path.join(REPO_ROOT, 'src', 'assets', 'data', 'cah_content.json')
VERSION_TAG = 'master-list-extract-2026-04-22'

# Left/right Set-name spelling reconciliation. Apply BEFORE filtering.
SET_NAME_NORMALIZE = {
    'CAH Everything Box Expanion':        'CAH Everything Box Expansion',
    'Dad Pack':                           'CAH Dad Pack',
    'House of Cards Pack':                'CAH House of Cards Pack',
    'Jack White Show Pack':               'CAH Jack White Show Pack',
    'Jew Pack/Chosen People Pack':        'CAH Jew Pack/Chosen People Pack',
    'Gen Con 2018 Midterm Election Pack': 'CAH Gen Con 2018 Midterm Election Pack',
    'CAH PAX East 2013 Promo Pack A':     'CAH Pax East 2013 Promo Pack A',
}


def normalize_set(raw):
    if raw is None:
        return None
    s = str(raw).strip()  # strip BOTH sides; covers the "trailing space" issue
    if not s:
        return None
    return SET_NAME_NORMALIZE.get(s, s)


def derive_pick(special_raw, text):
    special = '' if special_raw is None else str(special_raw).strip().upper()
    if special:
        m = re.search(r'PICK\s*(\d)', special)
        if m:
            return int(m.group(1))
    blanks = len(re.findall(r'_{2,}', text or ''))
    if blanks >= 3:
        return 3
    if blanks == 2:
        return 2
    return 1


def main():
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['Master Cards List']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    packs = defaultdict(lambda: {'prompts': [], 'responses': []})
    prompt_seen = defaultdict(set)
    response_seen = defaultdict(set)

    for row in rows[1:]:  # skip r1 header
        # LEFT half — prompts
        if len(row) >= 3:
            prompt_raw, special_raw, set_l_raw = row[0], row[1], row[2]
            set_l = normalize_set(set_l_raw)
            if set_l and set_l.startswith('CAH ') and prompt_raw and str(prompt_raw).strip():
                text = str(prompt_raw).strip()
                if text not in prompt_seen[set_l]:
                    prompt_seen[set_l].add(text)
                    packs[set_l]['prompts'].append({
                        'text': text,
                        'pick': derive_pick(special_raw, text),
                    })
        # RIGHT half — responses
        if len(row) >= 8:
            resp_raw, set_r_raw = row[6], row[7]
            set_r = normalize_set(set_r_raw)
            if set_r and set_r.startswith('CAH ') and resp_raw and str(resp_raw).strip():
                text = str(resp_raw).strip()
                if text not in response_seen[set_r]:
                    response_seen[set_r].add(text)
                    packs[set_r]['responses'].append({'text': text})

    # Drop zero/zero packs (defensive; shouldn't occur after CAH-prefix filter)
    filtered = {n: d for n, d in packs.items() if d['prompts'] or d['responses']}
    names = sorted(filtered)

    # Report
    print(f'\n{"Pack name":<55}  {"prompts":>7}  {"responses":>9}')
    print('-' * 76)
    for name in names:
        p = len(filtered[name]['prompts'])
        r = len(filtered[name]['responses'])
        print(f'{name:<55}  {p:>7}  {r:>9}')
    total_p = sum(len(d['prompts']) for d in filtered.values())
    total_r = sum(len(d['responses']) for d in filtered.values())
    print('-' * 76)
    print(f'{"TOTALS":<55}  {total_p:>7}  {total_r:>9}')
    print(f'\nPacks: {len(filtered)}')

    asym = []
    for name in names:
        p = len(filtered[name]['prompts'])
        r = len(filtered[name]['responses'])
        if (p > 0) != (r > 0):
            asym.append((name, p, r))
    print(f'\nPacks with only one side populated ({len(asym)}):')
    if not asym:
        print('  (none)')
    else:
        for name, p, r in asym:
            side = 'prompts-only' if p > 0 else 'responses-only'
            print(f'  {side:<15}  {name}  (p={p}, r={r})')

    # Write
    output = {
        'version': VERSION_TAG,
        'packs': {
            name: {
                'name': name,
                'prompts': filtered[name]['prompts'],
                'responses': filtered[name]['responses'],
            }
            for name in names
        },
    }
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    size = os.path.getsize(OUT_PATH)
    with open(OUT_PATH, 'rb') as f:
        md5 = hashlib.md5(f.read()).hexdigest()
    print(f'\nWrote {OUT_PATH}')
    print(f'  size: {size} bytes')
    print(f'  md5:  {md5}')


if __name__ == '__main__':
    main()
